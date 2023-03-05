import math
import random
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

a = ['100117A','100702A','111020A','120804A','140903A','141212A','150423A','150710A','180727A']
a1 = [[] for _ in range(len(a))]
a2 = [[] for _ in range(len(a))]
for r in range(0, len(a)):

    a1[r] = a[r]+'.xlsx'
    a2[r] = a[r]+'.txt'

def ope(a):
    wb = openpyxl.load_workbook(a) # 打开excel文件
    sheet = wb.active  # 正对表格

    n1 = sheet.max_row
    n2 = 11

    data = [[] for _ in range(n1)]
    list = [[] for _ in range(n2)]

    for column in range(1, n1+1):
        for row in range(1, n2+1):
            list[row-1] = str(sheet.cell(column, row).value)
        data[column-1]=list
        list = [[0] for _ in range(n2)]
    return data

def mainw(wjm, h):
    file_handle = open(wjm, mode='w')


    for j in range(0, len(h)):

        for i in range(0, 11):
            if i == 10:
                file_handle.writelines([h[j][i] + '\n'])
            else:
                file_handle.writelines([h[j][i]+' '])

    file_handle.close()

for g in range(0, len(a)):
    m = ope(a1[g])
    # print(m)

    n1 = len(m)
    n2=11
    da = [[] for _ in range(n1)]
    li = [[] for _ in range(n2)]

    for column in range(1, n1+1):
        for row in range(1, n2+1):
            li[row-1] = '0'
        da[column-1]=li
        li = [[] for _ in range(n2)]

    for i in range(0, len(m)):
        da[i][0] = m[i][0]
        for j in range(0, 11):
            if m[i][j] == '--' or m[i][j] == 'None':
                m[i][j] = '0'
        da[i][1] = m[i][1]
        da[i][3] = m[i][3]
        da[i][5] = m[i][5]
    hh = mainw(a2[g], da)






